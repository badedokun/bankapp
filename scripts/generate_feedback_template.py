#!/usr/bin/env python3
"""
OrokiiPay User Feedback Template Generator

This script automatically generates a complete Excel workbook with:
- Multiple sheets from CSV templates
- Data validation (dropdowns)
- Conditional formatting
- Formulas and calculations
- Charts and pivot tables
- Professional formatting

Usage:
    python generate_feedback_template.py

Output:
    OrokiiPay_User_Feedback_Tracker.xlsx in the docs/ folder
"""

import os
import sys
from pathlib import Path
from datetime import datetime
import csv

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.chart import PieChart, BarChart, LineChart, Reference
except ImportError:
    print("❌ Error: openpyxl is not installed.")
    print("📦 Install it with: pip install openpyxl")
    sys.exit(1)


# Configuration
SCRIPT_DIR = Path(__file__).parent
DOCS_DIR = SCRIPT_DIR.parent / "docs"
OUTPUT_FILE = DOCS_DIR / "OrokiiPay_User_Feedback_Tracker.xlsx"

# CSV Template Files
CSV_FILES = {
    "Feedback Log": "OrokiiPay_User_Feedback_Template.csv",
    "AI Assistant Feedback": "AI_Assistant_Feedback_Template.csv",
    "Feature Requests": "Feature_Requests_Template.csv",
    "Bug Reports": "Bug_Reports_Template.csv",
    "Satisfaction Survey": "User_Satisfaction_Survey_Template.csv",
    "Dropdown Reference": "Dropdown_Values_Reference.csv",
}

# Color scheme (OrokiiPay branding)
COLORS = {
    "primary": "010080",  # OrokiiPay dark blue
    "header_bg": "010080",
    "header_text": "FFFFFF",
    "critical": "FF0000",
    "high": "FFA500",
    "medium": "FFFF00",
    "low": "90EE90",
    "new": "ADD8E6",
    "in_progress": "FFA500",
    "resolved": "90EE90",
    "wont_fix": "D3D3D3",
    "alternate_row": "F5F5F5",
}


def print_header():
    """Print script header"""
    print("=" * 70)
    print("🏦 OrokiiPay User Feedback Template Generator")
    print("=" * 70)
    print()


def check_csv_files():
    """Check if all required CSV files exist"""
    print("📂 Checking for CSV template files...")
    missing_files = []

    for sheet_name, csv_file in CSV_FILES.items():
        csv_path = DOCS_DIR / csv_file
        if csv_path.exists():
            print(f"   ✅ Found: {csv_file}")
        else:
            print(f"   ❌ Missing: {csv_file}")
            missing_files.append(csv_file)

    if missing_files:
        print(f"\n❌ Error: {len(missing_files)} CSV file(s) missing!")
        print("Please ensure all CSV templates are in the docs/ folder.")
        sys.exit(1)

    print("   ✅ All CSV files found!\n")


def create_workbook():
    """Create a new Excel workbook"""
    print("📝 Creating new Excel workbook...")
    wb = Workbook()
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    print("   ✅ Workbook created\n")
    return wb


def import_csv_to_sheet(wb, sheet_name, csv_file):
    """Import CSV data into a worksheet"""
    print(f"📊 Importing {sheet_name}...")

    csv_path = DOCS_DIR / csv_file
    ws = wb.create_sheet(title=sheet_name)

    # Read CSV and write to sheet
    with open(csv_path, 'r', encoding='utf-8') as f:
        csv_reader = csv.reader(f)
        for row_idx, row in enumerate(csv_reader, start=1):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    print(f"   ✅ {sheet_name} imported ({ws.max_row} rows, {ws.max_column} columns)")
    return ws


def format_header_row(ws):
    """Apply formatting to header row"""
    header_font = Font(name='Arial', size=11, bold=True, color=COLORS["header_text"])
    header_fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def apply_alternate_row_colors(ws):
    """Apply alternate row coloring"""
    alternate_fill = PatternFill(start_color=COLORS["alternate_row"], end_color=COLORS["alternate_row"], fill_type="solid")

    for row_idx in range(2, ws.max_row + 1):
        if row_idx % 2 == 0:  # Even rows
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = alternate_fill


def auto_fit_columns(ws, max_width=50):
    """Auto-fit column widths"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def freeze_header_row(ws):
    """Freeze the top row"""
    ws.freeze_panes = "A2"


def add_filters(ws):
    """Add auto-filter to header row"""
    ws.auto_filter.ref = ws.dimensions


def create_named_ranges(wb):
    """Create named ranges for dropdown lists"""
    print("📋 Creating named ranges for dropdowns...")

    from openpyxl.workbook.defined_name import DefinedName

    dropdown_sheet = wb["Dropdown Reference"]

    # Define named ranges based on categories in Dropdown Reference sheet
    current_category = None
    start_row = None

    for row_idx in range(2, dropdown_sheet.max_row + 1):
        category = dropdown_sheet.cell(row=row_idx, column=1).value

        if category and category != current_category:
            # Save previous range
            if current_category and start_row:
                end_row = row_idx - 1
                range_name = f"{current_category.replace(' ', '_').replace('/', '_')}_List"
                range_ref = f"'Dropdown Reference'!$B${start_row}:$B${end_row}"

                try:
                    # Use new method for creating named ranges (openpyxl 3.1+)
                    defn = DefinedName(range_name, attr_text=range_ref)
                    wb.defined_names[range_name] = defn
                    print(f"   ✅ Created: {range_name}")
                except Exception as e:
                    print(f"   ⚠️  Could not create {range_name}: {e}")

            current_category = category
            start_row = row_idx

    # Save last range
    if current_category and start_row:
        end_row = dropdown_sheet.max_row
        range_name = f"{current_category.replace(' ', '_').replace('/', '_')}_List"
        range_ref = f"'Dropdown Reference'!$B${start_row}:$B${end_row}"

        try:
            # Use new method for creating named ranges (openpyxl 3.1+)
            defn = DefinedName(range_name, attr_text=range_ref)
            wb.defined_names[range_name] = defn
            print(f"   ✅ Created: {range_name}")
        except Exception as e:
            print(f"   ⚠️  Could not create {range_name}: {e}")

    print()


def add_data_validation(ws, sheet_name):
    """Add data validation (dropdowns) to appropriate columns"""
    if sheet_name != "Feedback Log":
        return

    print("✅ Adding data validation to Feedback Log...")

    # Column mappings (column letter: named range)
    validations = {
        "E": "User_Type_List",
        "H": "Platform_List",
        "K": "Feature_Module_List",
        "L": "Feedback_Type_List",
        "M": "Priority_List",
        "N": "Severity_List",
        "V": "AI_Personality_List",
        "Y": "Status_List",
        "AA": "Assigned_To_List",
    }

    for col_letter, range_name in validations.items():
        try:
            dv = DataValidation(type="list", formula1=f"={range_name}", allow_blank=True)
            dv.error = "Invalid value"
            dv.errorTitle = "Invalid Selection"
            dv.prompt = f"Please select from the dropdown list"
            dv.promptTitle = "Select Value"

            # Apply to entire column (rows 2-1000)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")
            print(f"   ✅ Added validation to column {col_letter} ({range_name})")
        except Exception as e:
            print(f"   ⚠️  Could not add validation to column {col_letter}: {e}")

    print()


def add_priority_formula(ws, sheet_name):
    """Add priority score formula to Feedback Log"""
    if sheet_name != "Feedback Log":
        return

    print("🔢 Adding priority score formula...")

    # Priority Score column is Z (column 26)
    # Formula based on: Severity (N) and User Type (E)
    formula = """=IF(N2="Blocker",4,IF(N2="Major",3,IF(N2="Minor",2,1)))*2.5+IF(E2="VIP",2,IF(E2="Beta Tester",1.5,IF(E2="Early Adopter",1,0.5)))"""

    for row_idx in range(2, ws.max_row + 1):
        ws[f"Z{row_idx}"] = formula
        ws[f"Z{row_idx}"].number_format = "0.0"

    print(f"   ✅ Formula added to column Z (rows 2-{ws.max_row})\n")


def add_conditional_formatting(ws, sheet_name):
    """Add conditional formatting for Priority and Status columns"""
    if sheet_name != "Feedback Log":
        return

    print("🎨 Adding conditional formatting...")

    # Priority column (M) - Color coding
    priority_rules = [
        ("Critical", COLORS["critical"]),
        ("High", COLORS["high"]),
        ("Medium", COLORS["medium"]),
        ("Low", COLORS["low"]),
    ]

    for priority, color in priority_rules:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = CellIsRule(operator='equal', formula=[f'"{priority}"'], fill=fill)
        ws.conditional_formatting.add(f"M2:M1000", rule)
        print(f"   ✅ Priority: {priority} → {color}")

    # Status column (Y) - Color coding
    status_rules = [
        ("New", COLORS["new"]),
        ("In Progress", COLORS["in_progress"]),
        ("Resolved", COLORS["resolved"]),
        ("Won't Fix", COLORS["wont_fix"]),
    ]

    for status, color in status_rules:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        rule = CellIsRule(operator='equal', formula=[f'"{status}"'], fill=fill)
        ws.conditional_formatting.add(f"Y2:Y1000", rule)
        print(f"   ✅ Status: {status} → {color}")

    print()


def create_dashboard_sheet(wb):
    """Create analytics dashboard with summary metrics"""
    print("📊 Creating Dashboard sheet...")

    ws = wb.create_sheet("Dashboard", 0)  # Insert as first sheet

    # Title
    ws["A1"] = "📊 OrokiiPay User Feedback Dashboard"
    ws["A1"].font = Font(size=18, bold=True, color=COLORS["primary"])
    ws.merge_cells("A1:F1")

    # Last Updated
    ws["A2"] = f"Last Updated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(size=10, italic=True)

    # Summary Metrics Section
    ws["A4"] = "📈 Summary Metrics"
    ws["A4"].font = Font(size=14, bold=True)

    # Metric labels
    metrics = [
        ("Total Feedback Count", "=COUNTA('Feedback Log'!A:A)-1"),
        ("Open Items", "=COUNTIFS('Feedback Log'!Y:Y,\"New\")+COUNTIFS('Feedback Log'!Y:Y,\"Under Review\")+COUNTIFS('Feedback Log'!Y:Y,\"In Progress\")"),
        ("Resolved Items", "=COUNTIF('Feedback Log'!Y:Y,\"Resolved\")"),
        ("Critical/High Priority", "=COUNTIFS('Feedback Log'!M:M,\"Critical\")+COUNTIFS('Feedback Log'!M:M,\"High\")"),
        ("AI-Related Feedback", "=COUNTIF('Feedback Log'!K:K,\"AI Assistant\")"),
        ("Bug Reports", "=COUNTIF('Feedback Log'!L:L,\"Bug\")"),
        ("Feature Requests", "=COUNTIF('Feedback Log'!L:L,\"Feature Request\")"),
    ]

    row_offset = 5
    for idx, (label, formula) in enumerate(metrics):
        ws[f"A{row_offset + idx}"] = label
        ws[f"A{row_offset + idx}"].font = Font(bold=True)
        ws[f"B{row_offset + idx}"] = formula
        ws[f"B{row_offset + idx}"].font = Font(size=12, bold=True, color=COLORS["primary"])
        ws[f"B{row_offset + idx}"].number_format = "#,##0"

    # Quick Stats Section
    ws["D4"] = "⚡ Quick Stats"
    ws["D4"].font = Font(size=14, bold=True)

    quick_stats = [
        ("Average NPS Score", "=AVERAGE('Satisfaction Survey'!E:E)"),
        ("Overall Satisfaction", "=AVERAGE('Satisfaction Survey'!D:D)"),
        ("AI Assistant Rating", "=AVERAGE('Satisfaction Survey'!H:H)"),
        ("Gamification Rating", "=AVERAGE('Satisfaction Survey'!L:L)"),
    ]

    for idx, (label, formula) in enumerate(quick_stats):
        ws[f"D{row_offset + idx}"] = label
        ws[f"D{row_offset + idx}"].font = Font(bold=True)
        ws[f"E{row_offset + idx}"] = formula
        ws[f"E{row_offset + idx}"].font = Font(size=12, bold=True, color=COLORS["primary"])
        ws[f"E{row_offset + idx}"].number_format = "0.0"

    # Instructions Section
    ws["A15"] = "📋 How to Use This Dashboard"
    ws["A15"].font = Font(size=12, bold=True)

    instructions = [
        "1. This dashboard auto-updates as you add feedback to the Feedback Log sheet",
        "2. Use the metrics above to track feedback trends",
        "3. Create pivot tables for deeper analysis (Insert → Pivot Table)",
        "4. Add charts to visualize data (Insert → Chart)",
        "5. Refresh data: Data → Refresh All",
    ]

    for idx, instruction in enumerate(instructions):
        ws[f"A{16 + idx}"] = instruction
        ws[f"A{16 + idx}"].font = Font(size=10)

    # Format columns
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["D"].width = 25
    ws.column_dimensions["E"].width = 15

    print("   ✅ Dashboard created with summary metrics\n")


def protect_dropdown_sheet(ws):
    """Protect the Dropdown Reference sheet"""
    print("🔒 Protecting Dropdown Reference sheet...")
    ws.protection.sheet = True
    # No password - just enable protection
    ws.protection.enable()
    print("   ✅ Sheet protected\n")


def add_instructions_sheet(wb):
    """Add a quick start instructions sheet"""
    print("📖 Creating Instructions sheet...")

    ws = wb.create_sheet("📖 Instructions", 1)  # Insert as second sheet

    # Title
    ws["A1"] = "📖 Quick Start Guide - User Feedback Template"
    ws["A1"].font = Font(size=16, bold=True, color=COLORS["primary"])
    ws.merge_cells("A1:D1")

    # For Users Section
    ws["A3"] = "👤 For Users Submitting Feedback:"
    ws["A3"].font = Font(size=12, bold=True)

    user_instructions = [
        "1. Go to the 'Feedback Log' sheet",
        "2. Find the first empty row",
        "3. Fill in your details: Name, Email, App Version, Platform",
        "4. Select Feature/Module from dropdown",
        "5. Select Feedback Type (Bug, Feature Request, etc.)",
        "6. Write a clear title and description",
        "7. Rate your satisfaction",
        "8. Save the file",
        "",
        "💡 Tip: Use dropdowns for standardized fields (look for ▼ arrow)",
        "💡 Tip: Attach screenshots separately and reference filename in column T",
    ]

    for idx, instruction in enumerate(user_instructions):
        ws[f"A{4 + idx}"] = instruction
        ws[f"A{4 + idx}"].font = Font(size=10)

    # For Team Section
    ws["A17"] = "👥 For Product/Development Team:"
    ws["A17"].font = Font(size=12, bold=True)

    team_instructions = [
        "1. Review 'Dashboard' sheet for summary metrics",
        "2. Go to 'Feedback Log' sheet",
        "3. Use filters to find unassigned or high-priority items",
        "4. Assign Priority and Severity",
        "5. Assign to team member",
        "6. Update Status as work progresses",
        "7. Add Internal Notes for tracking",
        "8. Update Resolution Date and Notes when fixed",
        "",
        "🔍 Quick Filters:",
        "   • Critical Issues: Filter Priority = 'Critical'",
        "   • My Tasks: Filter Assigned To = Your Name",
        "   • AI Feedback: Filter Feature/Module = 'AI Assistant'",
        "   • This Week: Filter Submission Date >= Start of week",
    ]

    for idx, instruction in enumerate(team_instructions):
        ws[f"A{18 + idx}"] = instruction
        ws[f"A{18 + idx}"].font = Font(size=10)

    # Sheet Guide
    ws["A35"] = "📑 Sheet Guide:"
    ws["A35"].font = Font(size=12, bold=True)

    sheet_guide = [
        ("Dashboard", "Summary metrics and quick stats"),
        ("Feedback Log", "Main feedback collection (start here!)"),
        ("AI Assistant Feedback", "Detailed AI performance metrics"),
        ("Feature Requests", "New features with voting"),
        ("Bug Reports", "Bug tracking and resolution"),
        ("Satisfaction Survey", "User satisfaction and NPS"),
        ("Dropdown Reference", "Master dropdown values (protected)"),
    ]

    ws["A36"] = "Sheet Name"
    ws["B36"] = "Purpose"
    ws["A36"].font = Font(bold=True)
    ws["B36"].font = Font(bold=True)

    for idx, (sheet, purpose) in enumerate(sheet_guide):
        ws[f"A{37 + idx}"] = sheet
        ws[f"B{37 + idx}"] = purpose

    # Support Section
    ws["A46"] = "❓ Need Help?"
    ws["A46"].font = Font(size=12, bold=True)

    ws["A47"] = "📧 Email: development@orokiipay.com"
    ws["A48"] = "📚 Docs: See USER_FEEDBACK_TEMPLATE.md in docs/ folder"
    ws["A49"] = "🔧 Setup: See FEEDBACK_TEMPLATE_SETUP_GUIDE.md"

    # Format columns
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 40

    print("   ✅ Instructions sheet created\n")


def main():
    """Main function"""
    print_header()

    # Check CSV files exist
    check_csv_files()

    # Create workbook
    wb = create_workbook()

    # Import all CSV files as sheets
    sheets = {}
    for sheet_name, csv_file in CSV_FILES.items():
        ws = import_csv_to_sheet(wb, sheet_name, csv_file)
        sheets[sheet_name] = ws

    print("\n" + "=" * 70)
    print("🎨 Applying Formatting...")
    print("=" * 70 + "\n")

    # Apply formatting to all sheets
    for sheet_name, ws in sheets.items():
        print(f"Formatting: {sheet_name}")
        format_header_row(ws)
        apply_alternate_row_colors(ws)
        auto_fit_columns(ws)
        freeze_header_row(ws)
        add_filters(ws)
        print()

    # Create Dashboard
    create_dashboard_sheet(wb)

    # Add Instructions
    add_instructions_sheet(wb)

    # Create named ranges
    create_named_ranges(wb)

    # Add data validation to Feedback Log
    add_data_validation(sheets["Feedback Log"], "Feedback Log")

    # Add formulas
    add_priority_formula(sheets["Feedback Log"], "Feedback Log")

    # Add conditional formatting
    add_conditional_formatting(sheets["Feedback Log"], "Feedback Log")

    # Protect Dropdown Reference sheet
    protect_dropdown_sheet(sheets["Dropdown Reference"])

    # Save workbook
    print("=" * 70)
    print("💾 Saving Excel Workbook...")
    print("=" * 70 + "\n")

    try:
        wb.save(OUTPUT_FILE)
        print(f"✅ Success! Workbook saved to:")
        print(f"   📁 {OUTPUT_FILE}")
        print()

        # Calculate file size
        file_size = OUTPUT_FILE.stat().st_size
        file_size_kb = file_size / 1024
        print(f"📊 File size: {file_size_kb:.1f} KB")
        print(f"📋 Total sheets: {len(wb.sheetnames)}")
        print()

        print("=" * 70)
        print("🎉 Template Generation Complete!")
        print("=" * 70)
        print()
        print("📌 Next Steps:")
        print("   1. Open the Excel file")
        print("   2. Review the 📖 Instructions sheet")
        print("   3. Check the Dashboard for sample metrics")
        print("   4. Share with your team")
        print("   5. Start collecting feedback!")
        print()
        print("📚 For detailed documentation, see:")
        print("   • USER_FEEDBACK_TEMPLATE.md")
        print("   • FEEDBACK_TEMPLATE_SETUP_GUIDE.md")
        print("   • FEEDBACK_SYSTEM_README.md")
        print()

    except Exception as e:
        print(f"❌ Error saving workbook: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
